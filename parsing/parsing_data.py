import openpyxl
from decorators import measure_time, timer


def get_headers(path: str)-> tuple[str]: 
    """
    Функция открывает файл, берет заголовки из таблицы и помещает в список
    
    Возвращает список headers
    """
    
    workbook = openpyxl.load_workbook(path)
    
    sheet = workbook.active
    
    # заголовки
    headers = []
    
    # проход по столбцам в таблице в первой строке
    for col in sheet.iter_cols(max_row=1, max_col=6):
        
        # проход по ячейкам в столбцах
        for cell in col:
            headers.append(cell.value)
            
    headers_tuple = tuple(headers)
    
    return headers_tuple


@measure_time
@timer
def parse_data_from_xl(path: str) -> list[dict]:
    """
    Функция открывает файл, проходит по всем строкам, помешает данные в строке в словарь, а затем в словари добавляются в общий список данных.
    
    Возвращает список словарей с ключами из заголовков, со значениями из соответсвтующих ячеек.
    """
    
    workbook = openpyxl.load_workbook(path)
    
    sheet = workbook.active

    # список для хранения данных
    data = []
    
    # список заголовков
    headers = get_headers(path)

    # проход по строкам в таблице
    for row in sheet.iter_rows():
        
        # словарь для записи данных в одной строке
        row_data = {}
        
        # проход по ячейкам в строке
        for cell in row:
            
            # по списку с заголовком
            for head in headers:
                
                # записываем значение ключа словаря из заголовка, а значение из ячейки
                row_data[head] = cell.value
                
        # добавляем словарь в общий список
        data.append(row_data)
            
    workbook.close()
    
    return data

if __name__ == '__main__':
    data = parse_data_from_xl('./data/track_data.xlsx')

    # for row in data:
    #     print(row)





